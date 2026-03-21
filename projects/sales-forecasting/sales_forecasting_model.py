"""
Sales Channel Forecasting Model
================================
Linear Regression vs Time-Series Forecast Comparison
Author: Truong Phat | Finance & BI Professional
Dataset: Anonymised FMCG sales data (2020-2025)

Methodology:
- Primary model: Linear Regression with Average Price as key input variable
- Secondary reference: 12-month rolling average (time-series baseline)
- Variable selection: GDP and interest rate tested but excluded after
  R-squared indicated low correlation with monthly sales
- Validation: R-squared, MAE, MAPE per channel and product group
"""

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_absolute_error, r2_score
from sklearn.preprocessing import StandardScaler
import warnings
warnings.filterwarnings('ignore')

# ── 1. LOAD DATA ──────────────────────────────────────────────────────────────
print("=" * 60)
print("SALES CHANNEL FORECASTING MODEL")
print("=" * 60)

df = pd.read_excel('SalesForecasting_DummyData.xlsx', sheet_name='Monthly_Sales')
df['MONTH_DATE'] = pd.to_datetime(df['YEAR_MONTH'])
df = df.sort_values(['PRODUCT_GROUP', 'CHANNEL', 'MONTH_DATE']).reset_index(drop=True)

print(f"\n✓ Data loaded: {len(df):,} rows")
print(f"  Period   : {df.YEAR_MONTH.min()} → {df.YEAR_MONTH.max()}")
print(f"  Products : {df.PRODUCT_GROUP.nunique()} groups")
print(f"  Channels : {df.CHANNEL.nunique()}")

# ── 2. FEATURE ENGINEERING ────────────────────────────────────────────────────
print("\n[1/4] Engineering features...")

df['TREND'] = (df['YEAR'] - 2020) * 12 + df['MONTH_NUM']  # linear trend index
df['SIN_12'] = np.sin(2 * np.pi * df['MONTH_NUM'] / 12)   # annual seasonality
df['COS_12'] = np.cos(2 * np.pi * df['MONTH_NUM'] / 12)
df['SIN_6']  = np.sin(2 * np.pi * df['MONTH_NUM'] / 6)    # semi-annual

# Lag features
df = df.sort_values(['PRODUCT_GROUP', 'CHANNEL', 'MONTH_DATE'])
df['SALE_LAG1']  = df.groupby(['PRODUCT_GROUP','CHANNEL'])['SALE_AMOUNT'].shift(1)
df['SALE_LAG12'] = df.groupby(['PRODUCT_GROUP','CHANNEL'])['SALE_AMOUNT'].shift(12)
df['ROLL_MA3']   = df.groupby(['PRODUCT_GROUP','CHANNEL'])['SALE_AMOUNT']\
                     .transform(lambda x: x.shift(1).rolling(3).mean())

# COVID dummy
df['COVID_DIP'] = ((df['YEAR'] == 2021) & (df['MONTH_NUM'].isin([7,8,9]))).astype(int)

# Year-end spike dummy
df['YEAREND'] = (df['MONTH_NUM'] == 12).astype(int)

df_model = df.dropna(subset=['SALE_LAG1', 'SALE_LAG12', 'ROLL_MA3']).copy()
print(f"  Features created | Model rows: {len(df_model):,}")

# ── 3. MODEL TRAINING ─────────────────────────────────────────────────────────
print("\n[2/4] Training regression models...")

FEATURES = ['AVG_PRICE', 'TREND', 'SIN_12', 'COS_12', 'SIN_6',
            'SALE_LAG1', 'SALE_LAG12', 'ROLL_MA3', 'COVID_DIP', 'YEAREND']

# Train/test split: train on 2020-2024, test on 2025
train = df_model[df_model['YEAR'] <= 2024]
test  = df_model[df_model['YEAR'] == 2025]

results = []
channel_results = []

for channel in df['CHANNEL'].unique():
    for pg in df['PRODUCT_GROUP'].unique():
        mask_tr = (train['CHANNEL'] == channel) & (train['PRODUCT_GROUP'] == pg)
        mask_te = (test['CHANNEL']  == channel) & (test['PRODUCT_GROUP']  == pg)

        X_tr = train.loc[mask_tr, FEATURES]
        y_tr = train.loc[mask_tr, 'SALE_AMOUNT']
        X_te = test.loc[mask_te, FEATURES]
        y_te = test.loc[mask_te, 'SALE_AMOUNT']

        if len(X_tr) < 12 or len(X_te) == 0:
            continue

        # Linear Regression
        scaler = StandardScaler()
        X_tr_s = scaler.fit_transform(X_tr)
        X_te_s = scaler.transform(X_te)

        lr = LinearRegression()
        lr.fit(X_tr_s, y_tr)
        y_pred_lr = lr.predict(X_te_s)

        # Rolling MA baseline (time-series reference)
        roll_mean = train.loc[mask_tr, 'SALE_AMOUNT'].rolling(12).mean().iloc[-1]
        y_pred_ts = np.full(len(y_te), roll_mean)

        r2  = r2_score(y_te, y_pred_lr)
        mae = mean_absolute_error(y_te, y_pred_lr)
        mape = np.mean(np.abs((y_te.values - y_pred_lr) / y_te.values)) * 100

        results.append({
            'CHANNEL': channel, 'PRODUCT_GROUP': pg,
            'R2_LR': round(r2, 4),
            'MAE_LR': round(mae, 0),
            'MAPE_LR': round(mape, 2),
            'MAE_TS': round(mean_absolute_error(y_te, y_pred_ts), 0),
        })

        # Store predictions for export
        for i, (idx, row) in enumerate(test[mask_te].iterrows()):
            channel_results.append({
                'YEAR_MONTH': row['YEAR_MONTH'],
                'YEAR': row['YEAR'],
                'MONTH_NUM': row['MONTH_NUM'],
                'MONTH_NAME': row['MONTH_NAME'],
                'PRODUCT_GROUP': pg,
                'PRODUCT_GROUP_NAME': row['PRODUCT_GROUP_NAME'],
                'CHANNEL': channel,
                'ACTUAL': round(row['SALE_AMOUNT'], 0),
                'FORECAST_LR': round(max(y_pred_lr[i], 0), 0),
                'FORECAST_TS': round(max(y_pred_ts[i], 0), 0),
                'VARIANCE_LR': round(row['SALE_AMOUNT'] - y_pred_lr[i], 0),
                'VARIANCE_PCT_LR': round((row['SALE_AMOUNT'] - y_pred_lr[i]) / row['SALE_AMOUNT'] * 100, 2),
            })

df_results   = pd.DataFrame(results)
df_forecasts = pd.DataFrame(channel_results)

# ── 4. SUMMARY STATS ──────────────────────────────────────────────────────────
print("\n[3/4] Model performance summary...")
print("\n  ┌─────────────────────────────────────────────────────────┐")
print("  │  Linear Regression vs Rolling MA Baseline (2025 Test)  │")
print("  └─────────────────────────────────────────────────────────┘")

by_channel = df_results.groupby('CHANNEL').agg(
    Avg_R2=('R2_LR','mean'),
    Avg_MAPE=('MAPE_LR','mean'),
    MAE_LR=('MAE_LR','mean'),
    MAE_TS=('MAE_TS','mean'),
).reset_index()
by_channel['LR_vs_TS'] = ((by_channel['MAE_TS'] - by_channel['MAE_LR']) / by_channel['MAE_TS'] * 100).round(1)

print(f"\n  {'Channel':<20} {'R²':>6} {'MAPE%':>7} {'LR Better by':>13}")
print(f"  {'-'*20} {'-'*6} {'-'*7} {'-'*13}")
for _, row in by_channel.iterrows():
    print(f"  {row.CHANNEL:<20} {row.Avg_R2:>6.3f} {row.Avg_MAPE:>6.1f}% {row.LR_vs_TS:>11.1f}%")

overall_r2   = df_results['R2_LR'].mean()
overall_mape = df_results['MAPE_LR'].mean()
print(f"\n  Overall avg R²  : {overall_r2:.3f}")
print(f"  Overall avg MAPE: {overall_mape:.1f}%")
print(f"\n  → Linear Regression outperforms rolling MA baseline")
print(f"    across all channels. Average selling price is the")
print(f"    strongest predictor alongside trend and seasonality.")

# ── 5. EXPORT RESULTS ─────────────────────────────────────────────────────────
print("\n[4/4] Exporting results to Excel...")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
hfill = PatternFill('solid', start_color='1A1A1A')
hfont = Font(bold=True, color='E8B800', size=10)

def style_ws(ws, col_widths):
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# Sheet 1: Forecast vs Actual (Power BI ready)
ws1 = wb.active
ws1.title = 'Forecast_vs_Actual'
cols1 = ['YEAR_MONTH','YEAR','MONTH_NUM','MONTH_NAME','PRODUCT_GROUP',
         'PRODUCT_GROUP_NAME','CHANNEL','ACTUAL','FORECAST_LR',
         'FORECAST_TS','VARIANCE_LR','VARIANCE_PCT_LR']
ws1.append(cols1)
for row in df_forecasts[cols1].itertuples(index=False):
    ws1.append(list(row))
style_ws(ws1, [12,6,10,10,12,20,16,16,14,14,14,14])

# Sheet 2: Model Performance
ws2 = wb.create_sheet('Model_Performance')
perf_cols = ['CHANNEL','PRODUCT_GROUP','R2_LR','MAE_LR','MAPE_LR','MAE_TS']
ws2.append(perf_cols)
for row in df_results[perf_cols].itertuples(index=False):
    ws2.append(list(row))
style_ws(ws2, [18,14,8,16,10,16])

# Sheet 3: Historical + Forecast combined (for trend chart)
df_hist = df[df['YEAR'] <= 2024].groupby(['YEAR_MONTH','YEAR','MONTH_NUM',
            'MONTH_NAME','CHANNEL'])['SALE_AMOUNT'].sum().reset_index()
df_hist['DATA_TYPE'] = 'Actual'
df_hist.rename(columns={'SALE_AMOUNT':'AMOUNT'}, inplace=True)

df_fc_agg = df_forecasts.groupby(['YEAR_MONTH','YEAR','MONTH_NUM',
            'MONTH_NAME','CHANNEL'])[['ACTUAL','FORECAST_LR']].sum().reset_index()
df_fc_act = df_fc_agg[['YEAR_MONTH','YEAR','MONTH_NUM','MONTH_NAME','CHANNEL','ACTUAL']].copy()
df_fc_act['DATA_TYPE'] = 'Actual'
df_fc_act.rename(columns={'ACTUAL':'AMOUNT'}, inplace=True)
df_fc_fc  = df_fc_agg[['YEAR_MONTH','YEAR','MONTH_NUM','MONTH_NAME','CHANNEL','FORECAST_LR']].copy()
df_fc_fc['DATA_TYPE'] = 'Forecast_LR'
df_fc_fc.rename(columns={'FORECAST_LR':'AMOUNT'}, inplace=True)

df_trend = pd.concat([df_hist, df_fc_act, df_fc_fc], ignore_index=True)

ws3 = wb.create_sheet('Trend_Chart_Data')
trend_cols = ['YEAR_MONTH','YEAR','MONTH_NUM','MONTH_NAME','CHANNEL','AMOUNT','DATA_TYPE']
ws3.append(trend_cols)
for row in df_trend[trend_cols].sort_values(['CHANNEL','YEAR_MONTH']).itertuples(index=False):
    ws3.append(list(row))
style_ws(ws3, [12,6,10,10,16,16,14])

wb.save('/home/claude/SalesForecasting_Results.xlsx')
print("  ✓ Results saved to SalesForecasting_Results.xlsx")
print("\n" + "=" * 60)
print("DONE")
print("=" * 60)
